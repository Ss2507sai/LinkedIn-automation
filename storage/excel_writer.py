"""Excel export for prospect results."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from src.logger import get_logger
from storage.results import COLUMNS, ProspectResult

logger = get_logger()


class ExcelWriter:
    """Append prospect results to an Excel workbook."""

    def __init__(self, path: Path) -> None:
        self.path = path

    def append(self, result: ProspectResult) -> None:
        """Append a single result row to the workbook."""
        if self.path.exists():
            workbook = load_workbook(self.path)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Results"
            sheet.append(COLUMNS)

        row = [result.to_row()[col] for col in COLUMNS]
        sheet.append(row)

        # Auto-size columns (approximate)
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            letter = get_column_letter(col_idx)
            sheet.column_dimensions[letter].width = min(max(len(col_name) + 2, 15), 50)

        workbook.save(self.path)
        logger.info("Excel saved: %s", self.path)
